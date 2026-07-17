import os
import sys
import json
import csv
import time
import threading
import urllib.parse
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError

# Set stdout encoding to UTF-8 for Windows console
sys.stdout.reconfigure(encoding='utf-8')

# Configurations
KEY_PATH = r"c:\Users\Quan Tran\Documents\GitHub\Quan-AI-Agent\Project-skills\01_Thanhthaimotor\Thanhthaimotor-references\bigquery_key.json"
CSV_PATH = r"c:\Users\Quan Tran\Documents\GitHub\Quan-AI-Agent\Project-skills\01_Thanhthaimotor\Thanhthaimotor-references\Thanhthaimotor_checkindex.csv"
CACHE_PATH = r"c:\Users\Quan Tran\Documents\GitHub\Quan-AI-Agent\Project-skills\01_Thanhthaimotor\Thanhthaimotor-references\inspect_cache.json"
OUTPUT_EXCEL_PATH = r"c:\Users\Quan Tran\Documents\GitHub\Quan-AI-Agent\Project-skills\01_Thanhthaimotor\Thanhthaimotor-references\[TTMT]ALL-KW-Status.xlsx"
DAILY_QUOTA_LIMIT = 6000
NUM_WORKERS = 4  # Concurrency level

# Global variables for thread safety
cache_lock = threading.Lock()
api_lock = threading.Lock()
properties_lock = threading.Lock()
thread_local = threading.local()
cache = {}
inspected_count = 0
exhausted_properties = set()

def get_property_url(url):
    """Determine the GSC property URL based on the URL path."""
    if "/kien-thuc-dien-co" in url:
        return "https://thanhthaimotor.com/kien-thuc-dien-co/"
    elif "/video/" in url:
        return "https://thanhthaimotor.com/video/"
    else:
        return "https://thanhthaimotor.com/"

def get_gsc_service(key_path):
    """Initialize GSC API client using service account credentials."""
    scopes = ['https://www.googleapis.com/auth/webmasters.readonly']
    credentials = service_account.Credentials.from_service_account_file(
        key_path, scopes=scopes
    )
    return build('searchconsole', 'v1', credentials=credentials)

def get_thread_service():
    """Get or initialize thread-local GSC service instance to ensure thread safety."""
    if not hasattr(thread_local, "service"):
        # Initialize service inside the thread
        thread_local.service = get_gsc_service(KEY_PATH)
    return thread_local.service

def inspect_url(service, site_url, url_to_inspect):
    """Call Google Search Console URL Inspection API for a URL."""
    try:
        # Rate limit protection: serialize GSC API calls slightly
        with api_lock:
            time.sleep(0.4) # spacing requests
            
        request = {
            'inspectionUrl': url_to_inspect,
            'siteUrl': site_url
        }
        response = service.urlInspection().index().inspect(body=request).execute()
        
        result = response.get('inspectionResult', {})
        status_result = result.get('indexStatusResult', {})
        verdict = status_result.get('verdict')
        coverage_state = status_result.get('coverageState', '')
        last_crawl_time = status_result.get('lastCrawlTime', '')
        
        # Determine status:
        # PASS verdict means Indexed.
        # NEUTRAL verdict means we must check if coverageState has "indexed" and NOT "not indexed"
        is_indexed = False
        if verdict == 'PASS':
            is_indexed = True
        elif verdict == 'NEUTRAL':
            state_lower = coverage_state.lower()
            if 'indexed' in state_lower and 'not indexed' not in state_lower:
                is_indexed = True
                
        status = 'index' if is_indexed else 'Noindex'
        return status, verdict, coverage_state, last_crawl_time
    except HttpError as e:
        return 'ERROR', 'HTTP_ERROR', str(e), ''
    except Exception as e:
        return 'ERROR', 'UNKNOWN_ERROR', str(e), ''

def load_cache():
    """Load caching of already inspected URLs to save quota."""
    if os.path.exists(CACHE_PATH):
        try:
            with open(CACHE_PATH, 'r', encoding='utf-8') as f:
                data = json.load(f)
                cleaned = {}
                for u, info in data.items():
                    status = info.get('status')
                    cov = info.get('coverage_state', '')
                    # If it was marked index but contains "not indexed", change to Noindex
                    if status == 'index' and 'not indexed' in cov.lower():
                        info['status'] = 'Noindex'
                    cleaned[u] = info
                return cleaned
        except Exception as e:
            print(f"⚠️ Warning: Failed to load cache file: {e}. Starting fresh.")
    return {}

def save_cache_file():
    """Save inspected URL results to cache file in a thread-safe way."""
    with cache_lock:
        try:
            with open(CACHE_PATH, 'w', encoding='utf-8') as f:
                json.dump(cache, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"❌ Error: Failed to save cache file: {e}")

def process_url(url, assigned_site_url, total_to_check):
    global inspected_count
    
    site_url = assigned_site_url
    
    with properties_lock:
        if site_url in exhausted_properties:
            return None
        
    try:
        service = get_thread_service()
    except Exception as e:
        print(f"\n❌ Error initializing service for thread: {e}")
        return None
        
    status, verdict, coverage_state, last_crawl_time = inspect_url(service, site_url, url)
    
    if status == 'ERROR' and ('quota' in coverage_state.lower() or '429' in coverage_state):
        with properties_lock:
            exhausted_properties.add(site_url)
            print(f"\n❌ Quota Limit or Rate Limit Reached for property: {site_url}! Skipping this property. Details: {coverage_state}")
        return None
        
    with cache_lock:
        cache[url] = {
            'status': status,
            'verdict': verdict,
            'coverage_state': coverage_state,
            'last_crawl_time': last_crawl_time,
            'check_time': time.strftime('%Y-%m-%d %H:%M:%S'),
            'inspected_under_property': site_url
        }
        inspected_count += 1
        current_count = inspected_count
        
    # Print progress status
    pct = (current_count / total_to_check) * 100
    print(f"[{current_count}/{total_to_check}] ({pct:.1f}%) Checked ({site_url}): {url} -> {status} ({verdict})")
    
    # Save cache incrementally
    if current_count % 10 == 0:
        save_cache_file()
        
    return url

def main():
    global cache, inspected_count
    print("🚀 GSC URL Inspection Tool Started (Thread-Safe Concurrent Mode)")
    
    # Read URLs from CSV file
    if not os.path.exists(CSV_PATH):
        print(f"❌ CSV File not found at {CSV_PATH}")
        return
        
    csv_rows = []
    try:
        with open(CSV_PATH, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            header = next(reader, None) # skip header
            for row in reader:
                if row and row[0].strip():
                    while len(row) < 4:
                        row.append('')
                    csv_rows.append(row)
    except Exception as e:
        print(f"❌ Error reading CSV file: {e}")
        return
    print(f"📋 Loaded {len(csv_rows)} URLs from CSV (excluding header).")
    
    # Prompt user for run mode
    choice = input("Bạn muốn chạy từ đầu hay tiếp tục? (tu dau / tiep tuc): ").strip().lower()
    
    is_from_start = choice in ('tu dau', 'từ đầu', 'tudau', '1')
    
    if is_from_start:
        print("🔄 Đã chọn chạy từ đầu. Đang reset cache và cập nhật trạng thái CSV...")
        cache = {}
        save_cache_file()
        
        for row in csv_rows:
            row[1] = 'Chưa check'
            row[2] = ''
            row[3] = ''
            
        urls = [row[0] for row in csv_rows]
        urls_to_inspect = list(urls)
        
        # Save CSV immediately
        update_csv(urls, cache)
    else:
        print("▶️ Đã chọn tiếp tục. Đang lọc các URL chưa check...")
        cache = load_cache()
        print(f"💾 Loaded cache. Already checked URLs in cache: {len(cache)}")
        
        urls = [row[0] for row in csv_rows]
        urls_to_inspect = []
        for row in csv_rows:
            url = row[0]
            status_csv = row[1].strip()
            
            # Chỉ lọc các URL đang có trạng thái "Chưa check" hoặc trống ở cột B
            if status_csv in ('Chưa check', '', None):
                if url in cache and cache[url].get('status') != 'ERROR' and cache[url].get('last_crawl_time'):
                    continue
                urls_to_inspect.append(url)
        
    print(f"🔍 URLs pending inspection: {len(urls_to_inspect)}")
    
    if len(urls_to_inspect) == 0:
        print("🎉 All URLs have already been inspected! Updating CSV and generating Excel file directly...")
        update_csv(urls, cache)
        generate_excel(urls, cache)
        return
        
    # Filter pending URLs into their natural groups
    blog_pending = []
    video_pending = []
    sanpham_pending = []
    
    for url in urls_to_inspect:
        if "/kien-thuc-dien-co" in url:
            blog_pending.append(url)
        elif "/video/" in url:
            video_pending.append(url)
        else:
            sanpham_pending.append(url)
            
    # Assign properties to pending URLs based on quota limitations and priority rules
    scheduled_targets = []  # List of tuples: (url, assigned_site_url)
    
    # 1. Allocate up to 2000 to Blog GSC property
    blog_assigned = blog_pending[:2000]
    blog_excess = blog_pending[2000:]
    for url in blog_assigned:
        scheduled_targets.append((url, "https://thanhthaimotor.com/kien-thuc-dien-co/"))
        
    # 2. Allocate up to 2000 to Video GSC property
    video_assigned = video_pending[:2000]
    video_excess = video_pending[2000:]
    for url in video_assigned:
        scheduled_targets.append((url, "https://thanhthaimotor.com/video/"))
        
    # 3. Allocate to Sản phẩm GSC property (limit 2000)
    sanpham_quota = 2000
    sanpham_assigned_urls = []
    
    # Ưu tiên 1: Quét list thừa của Blog và Video trên Sản phẩm
    excess_pool = blog_excess + video_excess
    for url in excess_pool:
        if len(sanpham_assigned_urls) < sanpham_quota:
            sanpham_assigned_urls.append(url)
            scheduled_targets.append((url, "https://thanhthaimotor.com/"))
            
    # Ưu tiên 2: Quét list Sản phẩm tự nhiên còn lại
    for url in sanpham_pending:
        if len(sanpham_assigned_urls) < sanpham_quota:
            sanpham_assigned_urls.append(url)
            scheduled_targets.append((url, "https://thanhthaimotor.com/"))
            
    print(f"📊 Scheduled GSC property routing summary:")
    print(f"   - To Blog Property: {len(blog_assigned)} URLs (Excess skipped: {len(blog_excess)})")
    print(f"   - To Video Property: {len(video_assigned)} URLs (Excess skipped: {len(video_excess)})")
    print(f"   - To Sản phẩm Property (Main): {len(sanpham_assigned_urls)} URLs "
          f"(from Blog/Video excess: {min(len(excess_pool), sanpham_quota)}, "
          f"from natural pending: {len(sanpham_assigned_urls) - min(len(excess_pool), sanpham_quota)})")
    print(f"   - Total Scheduled for Check in this Batch: {len(scheduled_targets)}")
    
    if len(scheduled_targets) == 0:
        print("🎉 No pending URLs assigned for check in this batch. Updating CSV and generating Excel file directly...")
        update_csv(urls, cache)
        generate_excel(urls, cache)
        return

    # Reset counter for this run
    inspected_count = 0
    start_time = time.time()
    
    # Run thread pool
    with ThreadPoolExecutor(max_workers=NUM_WORKERS) as executor:
        futures = {
            executor.submit(process_url, url, assigned_site_url, len(scheduled_targets)): (url, assigned_site_url)
            for url, assigned_site_url in scheduled_targets
        }
        
        for future in as_completed(futures):
            with properties_lock:
                if len(exhausted_properties) >= 3:
                    print("\n❌ All GSC properties have reached their quota limits! Stopping thread pool.")
                    # Cancel remaining futures
                    for f in futures:
                        f.cancel()
                    break
            try:
                future.result()
            except Exception as e:
                print(f"❌ Thread raised exception: {e}")
                
    save_cache_file()
    elapsed = time.time() - start_time
    print(f"\n✅ Finished batch of {inspected_count} URLs in {elapsed:.1f}s (Average: {elapsed/max(1, inspected_count):.2f}s per URL).")
    
    # Update CSV with new results
    update_csv(urls, cache)
    
    # Generate Excel
    generate_excel(urls, cache)

def generate_excel(all_urls, cache):
    """Generate the Excel file containing only the Noindex URLs from the cache."""
    print("📊 Generating Excel output file...")
    wb = Workbook()
    ws = wb.active
    ws.title = "Noindex URLs"
    ws.views.sheetView[0].showGridLines = True
    
    # Styles
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    noindex_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    noindex_font = Font(name="Calibri", size=11, color="C65911", bold=True)
    
    # Headers
    ws.append(["URL", "Trạng thái GSC", "Lý do (Coverage State)"])
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = left_align
        
    # Write only Noindex URLs
    row_idx = 2
    noindex_count = 0
    missing_count = 0
    
    for url in all_urls:
        if url in cache:
            info = cache[url]
            status = info.get('status')
            if status == 'Noindex':
                ws.cell(row=row_idx, column=1, value=url).alignment = left_align
                ws.cell(row=row_idx, column=2, value="Noindex").font = noindex_font
                ws.cell(row=row_idx, column=2).fill = noindex_fill
                ws.cell(row=row_idx, column=2).alignment = center_align
                ws.cell(row=row_idx, column=3, value=info.get('coverage_state')).alignment = left_align
                row_idx += 1
                noindex_count += 1
        else:
            missing_count += 1
            
    # Adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    ws.row_dimensions[1].height = 25
    wb.save(OUTPUT_EXCEL_PATH)
    
    print(f"💾 Saved Excel to {OUTPUT_EXCEL_PATH}")
    print(f"📊 Summary of CSV URLs ({len(all_urls)}):")
    print(f"   - Checked Noindex: {noindex_count}")
    print(f"   - Pending Checks: {missing_count}")
    
def get_property_display_name_fallback(url, cache_entry=None):
    if cache_entry and "inspected_under_property" in cache_entry:
        site_url = cache_entry["inspected_under_property"]
        if "kien-thuc-dien-co" in site_url:
            return "Blog"
        elif "video" in site_url:
            return "Video"
        else:
            return "Sản Phẩm"
    
    # Fallback to static matching
    if "/kien-thuc-dien-co" in url:
        return "Blog"
    elif "/video/" in url:
        return "Video"
    else:
        return "Sản Phẩm"

def update_csv(all_urls, cache):
    """Update the CSV file with URL, Index/Noindex status, Property Name, and Last Crawl Time."""
    print(f"📝 Updating CSV file at {CSV_PATH}...")
    try:
        with open(CSV_PATH, 'w', encoding='utf-8', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(["URLs check", "trạng thái", "Tên Property", "Thời gian cào"])
            for url in all_urls:
                if not url or url.upper() == "URL" or url == "URLs check":
                    continue
                
                # Determine status
                status = "Chưa check"
                prop_name = ""
                last_crawl = ""
                cache_entry = None
                if url in cache:
                    cache_entry = cache[url]
                    c_status = cache_entry.get('status')
                    if c_status == 'index':
                        status = 'Index'
                    elif c_status == 'Noindex':
                        status = 'Noindex'
                    
                    last_crawl = cache_entry.get('last_crawl_time', '')
                    if last_crawl:
                        # Clean date format: 2026-07-08T15:20:00Z -> 2026-07-08 15:20:00
                        last_crawl = last_crawl.replace('T', ' ').replace('Z', '')
                
                    # Determine property name only if it was checked
                    prop_name = get_property_display_name_fallback(url, cache_entry)
                    
                writer.writerow([url, status, prop_name, last_crawl])
        print(f"✅ CSV file successfully updated.")
    except Exception as e:
        print(f"❌ Error updating CSV file: {e}")

if __name__ == "__main__":
    main()
