import os
import sys
import json
import csv
import time
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError

# Set stdout encoding to UTF-8 for Windows console
sys.stdout.reconfigure(encoding='utf-8')

# Configurations
KEY_PATH = r"c:\Users\Quan Tran\Documents\GitHub\Quan-AI-Agent\Project-skills\04_Avakids\Avakids-references\gsc_key.json"
EXCEL_INPUT_PATH = r"c:\Users\Quan Tran\Documents\GitHub\Quan-AI-Agent\Project-skills\04_Avakids\Avakids-references\All_URL_Kids.xlsx"
CSV_PATH = r"c:\Users\Quan Tran\Documents\GitHub\Quan-AI-Agent\Project-skills\04_Avakids\Avakids-references\Avakids_checkindex.csv"
CACHE_PATH = r"c:\Users\Quan Tran\Documents\GitHub\Quan-AI-Agent\Project-skills\04_Avakids\Avakids-references\inspect_cache.json"
OUTPUT_EXCEL_PATH = r"c:\Users\Quan Tran\Documents\GitHub\Quan-AI-Agent\Project-skills\04_Avakids\Avakids-references\[AVAKIDS]ALL-KW-Status.xlsx"
DAILY_QUOTA_LIMIT = 4000
NUM_WORKERS = 40  # Concurrency level
REQUESTS_PER_SECOND = 8  # Safety limit for Google Search Console API rate limits (Max 600 QPM/10 QPS per project)

LINE_CHANNEL_ACCESS_TOKEN = "BMvkqGcs0dG4/+9iB0nrWuOJlmWivN+Rr9wTRYY8lFe9Ghqo8sOeY9aBrxfs8BcGh+jtGrZo9rriWsJBBLOfuvWJcY7fH9PwBQ+QIAkudR5EsCC4+u1/EDM2PRcTPRYatQzJM4S5AplFH51CItofXwdB04t89/1O/w1cDnyilFU="
LINE_TARGET_ID = "Cb83d530f2a25ad4fa63136fd8d9f00c9"

def set_console_title(title):
    """Set the terminal/console window title dynamically."""
    try:
        sys.stdout.write(f"\x1b]2;{title}\x07")
        sys.stdout.flush()
    except Exception:
        pass
    try:
        if os.name == 'nt':
            import ctypes
            ctypes.windll.kernel32.SetConsoleTitleW(title)
    except Exception:
        pass

class RateLimiter:
    """Thread-safe rate limiter to space the start time of API requests."""
    def __init__(self, requests_per_second):
        self.delay = 1.0 / requests_per_second
        self.last_time = time.time()
        self.lock = threading.Lock()

    def wait(self):
        with self.lock:
            current = time.time()
            elapsed = current - self.last_time
            if elapsed < self.delay:
                time.sleep(self.delay - elapsed)
                self.last_time = time.time()
            else:
                self.last_time = current

# Global variables for thread safety
cache_lock = threading.Lock()
rate_limiter = RateLimiter(REQUESTS_PER_SECOND)
properties_lock = threading.Lock()
thread_local = threading.local()
cache = {}
inspected_count = 0
exhausted_properties = set()

def get_property_url(url):
    """Determine the GSC property URL based on the URL path."""
    if "/me-va-be" in url:
        return "https://www.avakids.com/me-va-be/"
    else:
        return "https://www.avakids.com/"

def get_gsc_service(key_path):
    """Initialize GSC API client using service account credentials."""
    scopes = ['https://www.googleapis.com/auth/webmasters.readonly']
    credentials = service_account.Credentials.from_service_account_file(
        key_path, scopes=scopes
    )
    return build('searchconsole', 'v1', credentials=credentials)

def get_thread_service():
    """Get or initialize thread-local GSC service instance to ensure thread safety."""
    if not hasattr(thread_local, "service"):
        # Initialize service inside the thread
        thread_local.service = get_gsc_service(KEY_PATH)
    return thread_local.service

def inspect_url(service, site_url, url_to_inspect):
    """Call Google Search Console URL Inspection API for a URL."""
    try:
        # Rate limit protection: pace the start of GSC API requests globally
        rate_limiter.wait()
            
        request = {
            'inspectionUrl': url_to_inspect,
            'siteUrl': site_url
        }
        response = service.urlInspection().index().inspect(body=request).execute()
        
        result = response.get('inspectionResult', {})
        status_result = result.get('indexStatusResult', {})
        verdict = status_result.get('verdict')
        coverage_state = status_result.get('coverageState', '')
        last_crawl_time = status_result.get('lastCrawlTime', '')
        
        # Determine status:
        # PASS verdict means Indexed.
        # NEUTRAL verdict means we must check if coverageState has "indexed" and NOT "not indexed"
        is_indexed = False
        if verdict == 'PASS':
            is_indexed = True
        elif verdict == 'NEUTRAL':
            state_lower = coverage_state.lower()
            if 'indexed' in state_lower and 'not indexed' not in state_lower:
                is_indexed = True
                
        status = 'index' if is_indexed else 'Noindex'
        return status, verdict, coverage_state, last_crawl_time
    except HttpError as e:
        return 'ERROR', 'HTTP_ERROR', str(e), ''
    except Exception as e:
        return 'ERROR', 'UNKNOWN_ERROR', str(e), ''

def send_line_report(message):
    """Send automated report message to LINE group or user via LINE Messaging API."""
    if not LINE_CHANNEL_ACCESS_TOKEN or not LINE_TARGET_ID:
        print("ℹ️ LINE configuration is empty. Skipping LINE notification.")
        return
        
    import urllib.request
    import json
    
    url = "https://api.line.me/v2/bot/message/push"
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {LINE_CHANNEL_ACCESS_TOKEN}"
    }
    body = {
        "to": LINE_TARGET_ID,
        "messages": [
            {
                "type": "text",
                "text": message
            }
        ]
    }
    req = urllib.request.Request(
        url,
        data=json.dumps(body).encode('utf-8'),
        headers=headers,
        method='POST'
    )
    try:
        with urllib.request.urlopen(req) as response:
            if response.getcode() == 200:
                print("🔔 LINE report message sent successfully!")
            else:
                print(f"⚠️ LINE API returned status code: {response.getcode()}")
    except Exception as e:
        print(f"❌ Error sending LINE notification: {e}")

def load_cache():
    """Load caching of already inspected URLs to save quota."""
    if os.path.exists(CACHE_PATH):
        try:
            with open(CACHE_PATH, 'r', encoding='utf-8') as f:
                data = json.load(f)
                cleaned = {}
                for u, info in data.items():
                    status = info.get('status')
                    cov = info.get('coverage_state', '')
                    # If it was marked index but contains "not indexed", change to Noindex
                    if status == 'index' and 'not indexed' in cov.lower():
                        info['status'] = 'Noindex'
                    cleaned[u] = info
                return cleaned
        except Exception as e:
            print(f"⚠️ Warning: Failed to load cache file: {e}. Starting fresh.")
    return {}

def save_cache_file():
    """Save inspected URL results to cache file in a thread-safe way."""
    with cache_lock:
        try:
            with open(CACHE_PATH, 'w', encoding='utf-8') as f:
                json.dump(cache, f, ensure_ascii=False)
        except Exception as e:
            print(f"❌ Error: Failed to save cache file: {e}")

def process_url(url, assigned_site_url, total_to_check):
    global inspected_count
    
    site_url = assigned_site_url
    
    with properties_lock:
        if site_url in exhausted_properties:
            return None
        
    try:
        service = get_thread_service()
    except Exception as e:
        print(f"\n❌ Error initializing service for thread: {e}")
        return None
        
    status, verdict, coverage_state, last_crawl_time = inspect_url(service, site_url, url)
    
    if status == 'ERROR' and ('quota' in coverage_state.lower() or '429' in coverage_state):
        with properties_lock:
            exhausted_properties.add(site_url)
            print(f"\n❌ Quota Limit or Rate Limit Reached for property: {site_url}! Skipping this property. Details: {coverage_state}")
        return None
        
    with cache_lock:
        cache[url] = {
            'status': status,
            'verdict': verdict,
            'coverage_state': coverage_state,
            'last_crawl_time': last_crawl_time,
            'check_time': time.strftime('%Y-%m-%d %H:%M:%S'),
            'inspected_under_property': site_url
        }
        inspected_count += 1
        current_count = inspected_count
        
    # Print progress status
    pct = (current_count / total_to_check) * 100
    print(f"[{current_count}/{total_to_check}] ({pct:.1f}%) Checked ({site_url}): {url} -> {status} ({verdict})")
    set_console_title(f"[{current_count}/{total_to_check}] ({pct:.1f}%) GSC Inspect")
    
    # Save cache incrementally
    if current_count % 50 == 0:
        save_cache_file()
        
    return url

def load_urls_from_excel():
    """Load URLs and their status from Sheet1 and Sheet2 of the Excel file, keeping unique values in order."""
    if not os.path.exists(EXCEL_INPUT_PATH):
        print(f"❌ Excel Input File not found at {EXCEL_INPUT_PATH}")
        return [], {}
    
    try:
        wb = load_workbook(EXCEL_INPUT_PATH, read_only=True)
        urls = []
        seen = set()
        url_statuses = {}
        
        # Helper to read a sheet
        def read_sheet(sheet_name):
            if sheet_name not in wb.sheetnames:
                return
            sheet = wb[sheet_name]
            for row in sheet.iter_rows(values_only=True):
                if row and row[0]:
                    url = str(row[0]).strip()
                    if url and url.startswith("http"):
                        if url not in seen:
                            seen.add(url)
                            urls.append(url)
                        
                        status = None
                        if len(row) > 1 and row[1]:
                            status = str(row[1]).strip()
                        url_statuses[url] = status if status else "Chưa check"
        
        # Read Sheet1 then Sheet2
        read_sheet("Sheet1")
        read_sheet("Sheet2")
        wb.close()
        return urls, url_statuses
    except Exception as e:
        print(f"❌ Error reading Excel file: {e}")
        return [], {}

def main():
    global cache, inspected_count
    set_console_title("🚀 GSC URL Inspection Started - AVAKIDS")
    print("🚀 GSC URL Inspection Tool Started for AVAKIDS (Thread-Safe Concurrent Mode)")
    
    # Prompt the user for run mode
    run_mode = "continue"
    while True:
        choice = input("Bạn muốn chạy từ đầu hay tiếp tục? (tu dau / tiep tuc): ").strip().lower()
        if choice in ("từ đầu", "tu dau", "1", "tudau"):
            run_mode = "from_start"
            break
        elif choice in ("tiếp tục", "tiep tuc", "2", "tieptuc"):
            run_mode = "continue"
            break
        else:
            print("❌ Lựa chọn không hợp lệ. Vui lòng nhập 'từ đầu' hoặc 'tiếp tục'.")
            
    if run_mode == "from_start":
        reset_excel_column_b()
        
    # Load URLs from Excel
    excel_urls, url_statuses = load_urls_from_excel()
    if not excel_urls:
        print("❌ No URLs loaded. Exiting.")
        return
    print(f"📋 Loaded {len(excel_urls)} unique URLs from Excel file.")
    
    # Read existing status from CSV if it exists
    existing_status = {}
    if run_mode == "continue" and os.path.exists(CSV_PATH):
        try:
            with open(CSV_PATH, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                header = next(reader, None)  # skip header
                for row in reader:
                    if row and len(row) >= 4:
                        url = row[0].strip()
                        status = row[1].strip()
                        prop_name = row[2].strip()
                        last_crawl = row[3].strip()
                        existing_status[url] = (status, prop_name, last_crawl)
            print(f"📄 Loaded existing status for {len(existing_status)} URLs from CSV.")
        except Exception as e:
            print(f"⚠️ Warning: Error reading existing CSV: {e}. Reinitializing.")
            
    # Sync Excel URLs with existing CSV status
    final_urls = excel_urls
    
    # Load and clean cache
    if run_mode == "from_start":
        cache = {}
        save_cache_file()
        print("💾 Reset cache file as we are running from start.")
    else:
        cache = load_cache()
        print(f"💾 Loaded cache. Already checked URLs in cache: {len(cache)}")
        
        # Sync CSV records into cache if not present
        for url in final_urls:
            if url in existing_status and url not in cache:
                status, prop_name, last_crawl = existing_status[url]
                if status in ('Index', 'Noindex'):
                    cache_status = 'index' if status == 'Index' else 'Noindex'
                    cache[url] = {
                        'status': cache_status,
                        'verdict': 'PASS' if cache_status == 'index' else 'NEUTRAL',
                        'coverage_state': 'Indexed' if cache_status == 'index' else 'Not indexed',
                        'last_crawl_time': last_crawl,
                        'check_time': time.strftime('%Y-%m-%d %H:%M:%S'),
                        'inspected_under_property': get_property_url(url)
                    }

    # Filter out URLs to inspect
    urls_to_inspect = []
    for url in final_urls:
        if run_mode == "continue":
            # Only check URLs with 'Chưa check' status in Excel
            if url_statuses.get(url) != "Chưa check":
                continue
            # Standard cache check
            if url in cache and cache[url].get('status') != 'ERROR' and cache[url].get('last_crawl_time'):
                continue
        urls_to_inspect.append(url)
        
    print(f"🔍 URLs pending inspection: {len(urls_to_inspect)}")
    
    if len(urls_to_inspect) == 0:
        set_console_title("✅ Finished GSC Inspect - AVAKIDS")
        print("🎉 All URLs have already been inspected! Updating CSV, Excel input, and generating Excel file directly...")
        update_csv(final_urls, cache)
        update_excel_input(cache)
        generate_excel(final_urls, cache)
        
        # Send LINE report
        noindex_count = sum(1 for url in final_urls if url in cache and cache[url].get('status') == 'Noindex')
        index_count = sum(1 for url in final_urls if url in cache and cache[url].get('status') == 'index')
        report_message = (
            f"📊 BÁO CÁO KẾT QUẢ CHECK INDEX (AVAKIDS)\n"
            f"━━━━━━━━━━━━━━━━━━━\n"
            f"🎉 Tất cả URLs đã được check trước đó!\n"
            f"🔗 Tổng số URLs: {len(final_urls)}\n"
            f"🟢 Số URL đã Indexed: {index_count}\n"
            f"❌ Số URL lỗi Noindex: {noindex_count}\n"
            f"⏰ Thời gian: {time.strftime('%Y-%m-%d %H:%M:%S')}"
        )
        send_line_report(report_message)
        return
        
    # Filter pending URLs into their natural groups
    blog_pending = []
    sanpham_pending = []
    
    for url in urls_to_inspect:
        if "/me-va-be" in url:
            blog_pending.append(url)
        else:
            sanpham_pending.append(url)
            
    # Assign properties to pending URLs based on quota limitations and priority rules
    scheduled_targets = []  # List of tuples: (url, assigned_site_url)
    
    # 1. Allocate up to 2000 to Blog GSC property
    blog_assigned = blog_pending[:2000]
    blog_excess = blog_pending[2000:]
    for url in blog_assigned:
        scheduled_targets.append((url, "https://www.avakids.com/me-va-be/"))
        
    # 2. Allocate to Sản phẩm GSC property (limit 2000)
    sanpham_quota = 2000
    sanpham_assigned_urls = []
    
    # Ưu tiên 1: Quét list thừa của Blog trên Sản phẩm
    for url in blog_excess:
        if len(sanpham_assigned_urls) < sanpham_quota:
            sanpham_assigned_urls.append(url)
            scheduled_targets.append((url, "https://www.avakids.com/"))
            
    # Ưu tiên 2: Quét list Sản phẩm tự nhiên còn lại
    for url in sanpham_pending:
        if len(sanpham_assigned_urls) < sanpham_quota:
            sanpham_assigned_urls.append(url)
            scheduled_targets.append((url, "https://www.avakids.com/"))
            
    print(f"📊 Scheduled GSC property routing summary:")
    print(f"   - To Blog Property: {len(blog_assigned)} URLs (Excess routed to Main or skipped: {len(blog_excess)})")
    print(f"   - To Sản phẩm Property (Main): {len(sanpham_assigned_urls)} URLs "
          f"(from Blog excess: {min(len(blog_excess), sanpham_quota)}, "
          f"from natural pending: {len(sanpham_assigned_urls) - min(len(blog_excess), sanpham_quota)})")
    print(f"   - Total Scheduled for Check in this Batch: {len(scheduled_targets)}")
    
    if len(scheduled_targets) == 0:
        set_console_title("✅ Finished GSC Inspect - AVAKIDS")
        print("🎉 No pending URLs assigned for check in this batch. Updating CSV, Excel input, and generating Excel file directly...")
        update_csv(final_urls, cache)
        update_excel_input(cache)
        generate_excel(final_urls, cache)
        
        # Send LINE report
        noindex_count = sum(1 for url in final_urls if url in cache and cache[url].get('status') == 'Noindex')
        index_count = sum(1 for url in final_urls if url in cache and cache[url].get('status') == 'index')
        report_message = (
            f"📊 BÁO CÁO KẾT QUẢ CHECK INDEX (AVAKIDS)\n"
            f"━━━━━━━━━━━━━━━━━━━\n"
            f"🎉 Không có URL nào cần quét trong lượt này!\n"
            f"🔗 Tổng số URLs: {len(final_urls)}\n"
            f"🟢 Số URL đã Indexed: {index_count}\n"
            f"❌ Số URL lỗi Noindex: {noindex_count}\n"
            f"⏰ Thời gian: {time.strftime('%Y-%m-%d %H:%M:%S')}"
        )
        send_line_report(report_message)
        return

    # Reset counter for this run
    inspected_count = 0
    start_time = time.time()
    
    # Run thread pool
    with ThreadPoolExecutor(max_workers=NUM_WORKERS) as executor:
        futures = {
            executor.submit(process_url, url, assigned_site_url, len(scheduled_targets)): (url, assigned_site_url)
            for url, assigned_site_url in scheduled_targets
        }
        
        for future in as_completed(futures):
            with properties_lock:
                if len(exhausted_properties) >= 2:
                    print("\n❌ All GSC properties have reached their quota limits! Stopping thread pool.")
                    # Cancel remaining futures
                    for f in futures:
                        f.cancel()
                    break
            try:
                future.result()
            except Exception as e:
                print(f"❌ Thread raised exception: {e}")
                
    save_cache_file()
    elapsed = time.time() - start_time
    print(f"\n✅ Finished batch of {inspected_count} URLs in {elapsed:.1f}s (Average: {elapsed/max(1, inspected_count):.2f}s per URL).")
    set_console_title("✅ Finished GSC Inspect - AVAKIDS")
    
    # Update CSV with new results
    update_csv(final_urls, cache)
    
    # Update Excel Input
    update_excel_input(cache)
    
    # Generate Excel
    generate_excel(final_urls, cache)
    
    # Send LINE report
    noindex_count = sum(1 for url in final_urls if url in cache and cache[url].get('status') == 'Noindex')
    index_count = sum(1 for url in final_urls if url in cache and cache[url].get('status') == 'index')
    report_message = (
        f"📊 BÁO CÁO KẾT QUẢ CHECK INDEX (AVAKIDS)\n"
        f"━━━━━━━━━━━━━━━━━━━\n"
        f"✅ Hoàn thành quét URL GSC!\n"
        f"🔗 Tổng số URLs: {len(final_urls)}\n"
        f"🔍 Đã quét mới trong lượt này: {inspected_count}\n"
        f"🟢 Số URL đã Indexed: {index_count}\n"
        f"❌ Số URL lỗi Noindex: {noindex_count}\n"
        f"⏰ Thời gian: {time.strftime('%Y-%m-%d %H:%M:%S')}"
    )
    send_line_report(report_message)

def generate_excel(all_urls, cache):
    """Generate the Excel file containing only the Noindex URLs from the cache."""
    print("📊 Generating Excel output file...")
    wb = Workbook()
    ws = wb.active
    ws.title = "Noindex URLs"
    ws.views.sheetView[0].showGridLines = True
    
    # Styles
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    noindex_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    noindex_font = Font(name="Calibri", size=11, color="C65911", bold=True)
    
    # Headers
    ws.append(["URL", "Trạng thái GSC", "Lý do (Coverage State)"])
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = left_align
        
    # Write only Noindex URLs
    row_idx = 2
    noindex_count = 0
    missing_count = 0
    
    for url in all_urls:
        if url in cache:
            info = cache[url]
            status = info.get('status')
            if status == 'Noindex':
                ws.cell(row=row_idx, column=1, value=url).alignment = left_align
                ws.cell(row=row_idx, column=2, value="Noindex").font = noindex_font
                ws.cell(row=row_idx, column=2).fill = noindex_fill
                ws.cell(row=row_idx, column=2).alignment = center_align
                ws.cell(row=row_idx, column=3, value=info.get('coverage_state')).alignment = left_align
                row_idx += 1
                noindex_count += 1
        else:
            missing_count += 1
            
    # Adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    ws.row_dimensions[1].height = 25
    wb.save(OUTPUT_EXCEL_PATH)
    
    print(f"💾 Saved Excel to {OUTPUT_EXCEL_PATH}")
    print(f"📊 Summary of URLs ({len(all_urls)}):")
    print(f"   - Checked Noindex: {noindex_count}")
    print(f"   - Pending/Skipped/Indexed: {missing_count}")
    
def get_property_display_name_fallback(url, cache_entry=None):
    if cache_entry and "inspected_under_property" in cache_entry:
        site_url = cache_entry["inspected_under_property"]
        if "me-va-be" in site_url:
            return "Blog"
        else:
            return "Sản Phẩm"
    
    # Fallback to static matching
    if "/me-va-be" in url:
        return "Blog"
    else:
        return "Sản Phẩm"

def reset_excel_column_b():
    """Reset all status values in column B of Excel input file to 'Chưa check'."""
    if not os.path.exists(EXCEL_INPUT_PATH):
        print(f"❌ Excel Input File not found at {EXCEL_INPUT_PATH}")
        return
    print(f"📝 Resetting columns B, C, D in Excel input file at {EXCEL_INPUT_PATH}...")
    try:
        wb = load_workbook(EXCEL_INPUT_PATH)
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            for row in sheet.iter_rows(min_row=1):
                if not row or not row[0].value:
                    continue
                url = str(row[0].value).strip()
                # Check for header row
                if url == "URL" or url == "URLs check":
                    sheet.cell(row=row[0].row, column=2, value="trạng thái")
                    sheet.cell(row=row[0].row, column=3, value="Tên Property")
                    sheet.cell(row=row[0].row, column=4, value="Thời gian cào")
                    continue
                if not url.startswith("http"):
                    continue
                sheet.cell(row=row[0].row, column=2, value="Chưa check")
                sheet.cell(row=row[0].row, column=3, value="")
                sheet.cell(row=row[0].row, column=4, value="")
        wb.save(EXCEL_INPUT_PATH)
        print("✅ Excel input file successfully reset columns B, C, D.")
    except Exception as e:
        print(f"❌ Error resetting Excel input file: {e}")

def update_excel_input(cache):
    """Update the original All_URL_Kids.xlsx file columns B, C, D with check results."""
    if not os.path.exists(EXCEL_INPUT_PATH):
        print(f"❌ Excel Input File not found at {EXCEL_INPUT_PATH}")
        return
        
    print(f"📝 Updating original Excel input file at {EXCEL_INPUT_PATH}...")
    try:
        wb = load_workbook(EXCEL_INPUT_PATH)
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            for row in sheet.iter_rows(min_row=1):
                if not row or not row[0].value:
                    continue
                url = str(row[0].value).strip()
                
                # Check for header row
                if url == "URL" or url == "URLs check":
                    sheet.cell(row=row[0].row, column=2, value="trạng thái")
                    sheet.cell(row=row[0].row, column=3, value="Tên Property")
                    sheet.cell(row=row[0].row, column=4, value="Thời gian cào")
                    continue
                    
                if not url.startswith("http"):
                    continue
                    
                # Determine status
                status = "Chưa check"
                last_crawl = ""
                cache_entry = None
                if url in cache:
                    cache_entry = cache[url]
                    c_status = cache_entry.get('status')
                    if c_status == 'index':
                        status = 'Index'
                    elif c_status == 'Noindex':
                        status = 'Noindex'
                    
                    last_crawl = cache_entry.get('last_crawl_time', '')
                    if last_crawl:
                        last_crawl = last_crawl.replace('T', ' ').replace('Z', '')
                
                # Determine property name
                prop_name = get_property_display_name_fallback(url, cache_entry)
                
                sheet.cell(row=row[0].row, column=2, value=status)
                sheet.cell(row=row[0].row, column=3, value=prop_name)
                sheet.cell(row=row[0].row, column=4, value=last_crawl)
                
        wb.save(EXCEL_INPUT_PATH)
        print(f"✅ Excel input file successfully updated.")
    except Exception as e:
        print(f"❌ Error updating Excel input file: {e}")

def update_csv(all_urls, cache):
    """Update the CSV file with URL, Index/Noindex status, Property Name, and Last Crawl Time."""
    print(f"📝 Updating CSV file at {CSV_PATH}...")
    try:
        with open(CSV_PATH, 'w', encoding='utf-8', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(["URLs check", "trạng thái", "Tên Property", "Thời gian cào"])
            for url in all_urls:
                if not url or url.upper() == "URL" or url == "URLs check":
                    continue
                
                # Determine status
                status = "Chưa check"
                last_crawl = ""
                cache_entry = None
                if url in cache:
                    cache_entry = cache[url]
                    c_status = cache_entry.get('status')
                    if c_status == 'index':
                        status = 'Index'
                    elif c_status == 'Noindex':
                        status = 'Noindex'
                    
                    last_crawl = cache_entry.get('last_crawl_time', '')
                    if last_crawl:
                        # Clean date format: 2026-07-08T15:20:00Z -> 2026-07-08 15:20:00
                        last_crawl = last_crawl.replace('T', ' ').replace('Z', '')
                
                # Determine property name
                prop_name = get_property_display_name_fallback(url, cache_entry)
                    
                writer.writerow([url, status, prop_name, last_crawl])
        print(f"✅ CSV file successfully updated.")
    except Exception as e:
        print(f"❌ Error updating CSV file: {e}")

if __name__ == "__main__":
    main()
